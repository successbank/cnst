import openpyxl
import json
import os

def extract_product_data(excel_file, output_json):
    """Excel 파일에서 제품 데이터 추출"""
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        products = []
        
        # 헤더 행 찾기 (보통 '규격'과 '단위중량' 텍스트가 있는 행)
        header_row = None
        spec_col = None
        weight_col = None
        material_col = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and any('규격' in str(cell) if cell else False for cell in row):
                header_row = row_idx
                # 규격과 단위중량 열 찾기
                for col_idx, cell in enumerate(row):
                    if cell and '규격' in str(cell):
                        spec_col = col_idx
                    elif cell and '단위중량' in str(cell):
                        weight_col = col_idx
                    elif cell and '재질' in str(cell):
                        material_col = col_idx
                break
        
        if spec_col is None or weight_col is None:
            print(f"규격 또는 단위중량 열을 찾을 수 없음: {excel_file}")
            return 0
        
        # 데이터 추출
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            if row[spec_col] is None:  # 규격이 비어있으면 종료
                continue
                
            specification = str(row[spec_col]).strip() if row[spec_col] else None
            unit_weight = row[weight_col] if weight_col < len(row) else None
            material = row[material_col] if material_col and material_col < len(row) else None
            
            if specification and unit_weight is not None:
                try:
                    # 단중을 float로 변환
                    unit_weight = float(unit_weight)
                    product_data = {
                        'specification': specification,
                        'unit_weight': unit_weight
                    }
                    if material:
                        product_data['material'] = str(material).strip()
                    
                    products.append(product_data)
                except (ValueError, TypeError):
                    print(f"단중 변환 오류: {specification} - {unit_weight}")
        
        # JSON 파일로 저장
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(products, f, ensure_ascii=False, indent=2)
        
        print(f"{excel_file} 처리 완료: {len(products)}개 제품")
        return len(products)
        
    except Exception as e:
        print(f"오류 발생 ({excel_file}): {str(e)}")
        return 0

# 파일 목록
files = [
    ('114/8/BS파이프.xlsx', '114/8/bs_pipe_data.json'),
    ('114/8/KS파이프.xlsx', '114/8/ks_pipe_data.json'),
    ('114/8/강관파일.xlsx', '114/8/steel_pipe_pile_data.json'),
    ('114/8/구조관.xlsx', '114/8/structural_pipe_data.json'),
    ('114/8/단관비계.xlsx', '114/8/scaffold_pipe_data.json'),
    ('114/8/복공판.xlsx', '114/8/temporary_deck_data.json'),
    ('114/8/압력배관.xlsx', '114/8/pressure_pipe_data.json'),
    ('114/8/전선관.xlsx', '114/8/conduit_pipe_data.json')
]

total_count = 0
for excel_file, json_file in files:
    count = extract_product_data(excel_file, json_file)
    total_count += count

print(f"\n전체 처리 완료: 총 {total_count}개 제품")