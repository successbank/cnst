import openpyxl

# BS파이프 파일 구조 확인
excel_file = '114/8/BS파이프.xlsx'
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb.active

print(f"파일: {excel_file}")
print(f"시트명: {ws.title}")
print(f"최대 행: {ws.max_row}, 최대 열: {ws.max_column}")
print("\n처음 10행의 데이터:")

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
    print(f"행 {i+1}: {row}")
    
# 데이터가 있는 첫 번째 행 찾기
print("\n데이터가 있는 행 찾기:")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if any(cell is not None for cell in row):
        print(f"행 {i+1}: {row}")
        if i > 20:  # 처음 20행만 확인
            break